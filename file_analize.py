import os
import openpyxl as xl
from openpyxl.utils import get_column_letter
from copy import copy

def source_sheets(path) -> list:
    """
    Возвращает названия, кол-во строк и столбцов в листах
    из excel файла, находящегося по переданному пути

    :return: Список с названиями листов и количеством строк в таблице 
    :rtype: list
    """

    res = []
    wb = xl.load_workbook(path)
    for sheet in wb.sheetnames:
        res.append([sheet, wb[sheet].max_column, wb[sheet].max_row])
    return res

def lsr_documents(path) -> list:
    """
    Возвращает список со списками [код, расширенное название, полный путь к файлу]

    :return: Полная информация о всех файлах в директории с ЛСР 
    :rtype: list
    """
    res = []
    for file in os.listdir(path):
        if os.path.isfile(os.path.join(path, file)) and file.endswith('.xlsx'):
            res.append([file.split(' ')[0], (' '.join(file.split(' ')[1:])[:-5]), os.path.join(path, file)])
    return res


def create_error_table_header(Error_Worksheet, Main_Sheet) -> int:
    """
    Создает шапку таблицы для ошибок

    :return: None
    """

    PASTED_ROWS = 5

    for merge in list(Error_Worksheet.merged_cells):
        Error_Worksheet.unmerge_cells(range_string=str(merge))
    
    Error_Worksheet["A1"].value = "Ненайденные позиции из ЛСР"
    Error_Worksheet["A3"].value = "№ п/п"
    Error_Worksheet["B3"].value = "№ в ЛСР"
    Error_Worksheet["C3"].value = "Наименование работ"
    Error_Worksheet["D3"].value = "Ед.изм."
    Error_Worksheet["E3"].value = "Кол-во"
    Error_Worksheet["F3"].value = "Тип ошибки"

    for i, row in enumerate(Error_Worksheet["A1:F4"]):
        for j, cell in enumerate(row):
            if i==2: Error_Worksheet.cell(row=i+2, column=j+1).value = j+1
            Error_Worksheet.cell(row=i+1, column=j+1).font = copy(Main_Sheet["A1"].font)
            Error_Worksheet.cell(row=i+1, column=j+1).border = copy(Main_Sheet["A1"].border)
            Error_Worksheet.cell(row=i+1, column=j+1).fill = copy(Main_Sheet["A1"].fill)
            Error_Worksheet.cell(row=i+1, column=j+1).alignment = copy(Main_Sheet["A1"].alignment)

    for i, row in enumerate(Main_Sheet["A1:AV4"]):
        for j, cell in enumerate(row):
            if i==3: Error_Worksheet.cell(row=i+1, column=j+7, value=int(cell.value)+6)
            else: Error_Worksheet.cell(row=i+1, column=j+7, value=cell.value)
            Error_Worksheet.cell(row=i+1, column=j+7).font = copy(cell.font)
            Error_Worksheet.cell(row=i+1, column=j+7).border = copy(cell.border)
            Error_Worksheet.cell(row=i+1, column=j+7).fill = copy(cell.fill)
            Error_Worksheet.cell(row=i+1, column=j+7).number_format = copy(cell.number_format)
            Error_Worksheet.cell(row=i+1, column=j+7).protection = copy(cell.protection)
            Error_Worksheet.cell(row=i+1, column=j+7).alignment = copy(cell.alignment)
    
    for elem in Main_Sheet.merged_cells:
            elem=str(elem).split(':')
            elem[0] = list(xl.utils.cell.coordinate_to_tuple(elem[0]))
            elem[1] = list(xl.utils.cell.coordinate_to_tuple(elem[1]))
            Error_Worksheet.merge_cells(start_row=elem[0][0],
                                  start_column=elem[0][1]+6,
                                  end_row=elem[1][0],
                                  end_column=elem[1][1]+6)  
            
    Error_Worksheet.merge_cells("A1:F2")

    for i, width in enumerate([5, 40, 50, 13, 7, 30]): Error_Worksheet.column_dimensions[get_column_letter(i+1)].width = width
    for column in range(xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string("A1")[0]), xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string("AV4")[0])+1):
        Error_Worksheet.column_dimensions[get_column_letter(column+6)].width = int(Main_Sheet.column_dimensions[get_column_letter(column)].width)
    for i, height in enumerate([18, 36, 16]): Error_Worksheet.row_dimensions[i+1].height = height
    
    return PASTED_ROWS

def add_error_name_header(Error_Worksheet, Main_Sheet, sheet, PASTED_ROWS) -> int:
    """
    Создает заголовок для нового листа

    :return: None
    """

    Error_Worksheet[f"A{PASTED_ROWS}"].value = sheet
    Error_Worksheet[f"A{PASTED_ROWS}"].font = copy(Main_Sheet["A1"].font)
    Error_Worksheet[f"A{PASTED_ROWS}"].border = copy(Main_Sheet["A1"].border)
    Error_Worksheet[f"A{PASTED_ROWS}"].fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=xl.styles.colors.Color(rgb='bee5b3'))
    Error_Worksheet[f"A{PASTED_ROWS}"].alignment = copy(Main_Sheet["A1"].alignment)
    Error_Worksheet.merge_cells(f"A{PASTED_ROWS}:F{PASTED_ROWS}")
    
    return PASTED_ROWS+1

def add_error_row(Error_Worksheet, Main_Sheet, PASTED_ROWS: int, data, error_type: int, paste_data=None, source_sheet=None, source_index=None) -> None:
    """
    Создает новую строку с записью об ошибке

    :return: None
    """
    error_encoding = {
        101: ['Ресурс не найден', 'FF9194'],
        103: ['Неполные данные', 'ffe994'],
        104: ['Не совпадает ед. изм.', 'b9d7f0'], 
    }

    ERROR_COLUMN = 'F'

    for i, cell in enumerate(data):
        if i!=0:
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1).value = copy(cell.value)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1).font = copy(cell.font)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1).border = copy(cell.border)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1).fill = copy(cell.fill)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1).alignment = copy(cell.alignment)
    
    Error_Worksheet.cell(row=PASTED_ROWS, column=xl.utils.cell.column_index_from_string(ERROR_COLUMN)).value = f'Код ошибки: {error_type} : {error_encoding[error_type][0]}'
    Error_Worksheet.cell(row=PASTED_ROWS, column=xl.utils.cell.column_index_from_string(ERROR_COLUMN)).font = copy(Main_Sheet["A1"].font)
    Error_Worksheet.cell(row=PASTED_ROWS, column=xl.utils.cell.column_index_from_string(ERROR_COLUMN)).border = copy(Main_Sheet["A1"].border)
    Error_Worksheet.cell(row=PASTED_ROWS, column=xl.utils.cell.column_index_from_string(ERROR_COLUMN)).fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=xl.styles.colors.Color(rgb=error_encoding[error_type][1]))
    Error_Worksheet.cell(row=PASTED_ROWS, column=xl.utils.cell.column_index_from_string(ERROR_COLUMN)).alignment = copy(Main_Sheet["A1"].alignment)
    if source_sheet!= None and source_index!= None:
        Error_Worksheet.cell(row=PASTED_ROWS, column=xl.utils.cell.column_index_from_string(ERROR_COLUMN)).hyperlink = f"#'{source_sheet}'!A{source_index}"

    if paste_data:
        for i, cell in enumerate(paste_data):
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1+xl.utils.cell.column_index_from_string(ERROR_COLUMN)).value = copy(cell.value)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1+xl.utils.cell.column_index_from_string(ERROR_COLUMN)).font = copy(cell.font)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1+xl.utils.cell.column_index_from_string(ERROR_COLUMN)).border = copy(cell.border)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1+xl.utils.cell.column_index_from_string(ERROR_COLUMN)).fill = copy(cell.fill)
            Error_Worksheet.cell(row=PASTED_ROWS, column=i+1+xl.utils.cell.column_index_from_string(ERROR_COLUMN)).alignment = copy(cell.alignment)

    return PASTED_ROWS+1


def create_hyperlinks_table_header(Hyperlinks_Worksheet, Main_Sheet, source_sheets) -> int:

    pasted_rows=3 # кол-во служебных строк, которые есть по умолчанию

    Hyperlinks_Worksheet["A1"].value = "Гиперссылки на листы"
    Hyperlinks_Worksheet.merge_cells("A1:B1")

    Hyperlinks_Worksheet["A2"].value = "Листы с данными"
    Hyperlinks_Worksheet["A2"].fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=xl.styles.colors.Color(rgb='bee5b3'))
    Hyperlinks_Worksheet.merge_cells("A2:B2")

    for i, sheetname in enumerate(source_sheets):
        Hyperlinks_Worksheet[f"A{i+pasted_rows}"].value = sheetname
        Hyperlinks_Worksheet[f"A{i+pasted_rows}"].hyperlink = f"#'{sheetname}'!A1"
    pasted_rows+=len(source_sheets)
    
    Hyperlinks_Worksheet[f"A{pasted_rows}"].value = "Листы с ЛСР"
    Hyperlinks_Worksheet[f"A{pasted_rows}"].fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=xl.styles.colors.Color(rgb='bee5b3'))
    Hyperlinks_Worksheet.merge_cells(f"A{pasted_rows}:B{pasted_rows}")
    pasted_rows+=1

    Hyperlinks_Worksheet[f"A{pasted_rows}"].value = "Код"
    Hyperlinks_Worksheet[f"A{pasted_rows}"].fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=xl.styles.colors.Color(rgb='bee5b3'))
    Hyperlinks_Worksheet[f"B{pasted_rows}"].value = "Полное название источника"
    Hyperlinks_Worksheet[f"B{pasted_rows}"].fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=xl.styles.colors.Color(rgb='bee5b3'))

    Hyperlinks_Worksheet.column_dimensions['A'].width = 30
    Hyperlinks_Worksheet.column_dimensions['B'].width = 60

    for row in Hyperlinks_Worksheet[f"A1:B{pasted_rows}"]:
        for cell in row:
            cell.font = copy(Main_Sheet["A1"].font)
            cell.border = copy(Main_Sheet["A1"].border)
            cell.alignment = copy(Main_Sheet["A1"].alignment)
    
    return pasted_rows

def add_hyperlink_row(Hyperlinks_Worksheet, Main_Sheet, sheetname: str, full_name:str, LAST_ROW: int) -> None:
    
    Hyperlinks_Worksheet[f"A{LAST_ROW}"].value = sheetname
    Hyperlinks_Worksheet[f"A{LAST_ROW}"].hyperlink = f"#'{sheetname}'!A1"

    Hyperlinks_Worksheet[f"B{LAST_ROW}"].value = full_name
    for cell in Hyperlinks_Worksheet[LAST_ROW]:
        cell.font = copy(Main_Sheet["A1"].font)
        cell.border = copy(Main_Sheet["A1"].border)
        cell.alignment = copy(Main_Sheet["A1"].alignment)

    return None


def add_lsr_sheet(RES_FILE, lsr_document, Hyperlinks_Worksheet, Errors_Worksheet, source_sheets, pasted_links: int, pasted_errors: int) -> xl.workbook.workbook.Workbook:

    Main_Sheet = RES_FILE[source_sheets[0]]
    lsr_workbook_for_copy = xl.load_workbook(lsr_document[2])
    lsr_for_copy = lsr_workbook_for_copy[lsr_workbook_for_copy.sheetnames[0]]
    if lsr_document[0] not in RES_FILE.sheetnames:
        lsr_sheet = RES_FILE.create_sheet(lsr_document[0])
    else:
        lsr_sheet = RES_FILE.create_sheet(lsr_document[0]+" Новый")

    START_LSR_COLUMN = 'A'    # начало области копирования документа ЛСР (столбец)  
    START_LSR_ROW = '1'       # начало области копирования документа ЛСР (строка)  

    START_SOURCE_COLUMN = 'A'    # начало области копирования документа источкина (столбец)  
    START_SOURCE_ROW = '1'       # начало области копирования документа источника (строка)  

    END_LSR_COLUMN = 'A'      # последний столбец с информацией в документе ЛСР
    END_LSR_ROW = '1'         # последняя строка с информацией в документе ЛСР

    END_SOURCE_COLUMN = 'A'      # последний столбец с информацией в документе источника
    END_SOURCE_ROW = '1'         # последняя строка с информацией в документе источника

    LSR_NAME_COLUMN = 'C'     # столбец с названиями материалов в ЛСР
    LSR_UNIT_COLUMN = 'D'     # столбец с единицами измерений в ЛСР
    LSR_NUMS_ROW = '5'        # строка с индивидуальной нумерацией столбцов в ЛСР

    SOURCE_NAME_COLUMN = 'C'     # столбец с названиями материалов в источнике
    SOURCE_UNIT_COLUMN = 'E'     # столбец с единицами измерений в источнике
    SOURCE_NUMS_ROW = '4'        # строка с индивидуальной нумерацией столбцов в источнике

    # Определяет последнюю строку и столбец с данными для копирования из документа ЛСР
    for i, cell in enumerate(lsr_for_copy[LSR_NAME_COLUMN]): 
        if str(cell.value)!='None': END_LSR_ROW=str(i+1)
    for i, cell in enumerate(lsr_for_copy[int(LSR_NUMS_ROW)]): 
        if str(cell.value)!='None': END_LSR_COLUMN=xl.utils.cell.get_column_letter(i+1)
    
    # Определяет последнюю строку и столбец с данными для взаимодействия с источником
    for i, cell in enumerate(Main_Sheet[SOURCE_NAME_COLUMN]): 
        if str(cell.value)!='None': END_SOURCE_ROW=str(i+1)
    for i, cell in enumerate(Main_Sheet[int(SOURCE_NUMS_ROW)]): 
        if str(cell.value)!='None': END_SOURCE_COLUMN=xl.utils.cell.get_column_letter(i+1)

    END_LSR_INFO_COLUMN = 'E' # последний столбец с нужными данными, остальные за ним сжимаются
    END_LSR_INFO_ROW = str(int(END_LSR_ROW)-4)

    # Вычисление сдвигов заголовков для выравнивания
    LSR_SHIFT = 0
    SOURCE_SHIFT = 0

    if int(LSR_NUMS_ROW)>int(SOURCE_NUMS_ROW): SOURCE_SHIFT=int(LSR_NUMS_ROW)-int(SOURCE_NUMS_ROW)
    if int(LSR_NUMS_ROW)<int(SOURCE_NUMS_ROW): LSR_SHIFT=int(SOURCE_NUMS_ROW)-int(LSR_NUMS_ROW)

    # Копирование нужных ячеек из документа ЛСР и источника данных
    for i, row in enumerate(lsr_for_copy[f'{START_LSR_COLUMN}{START_LSR_ROW}:{END_LSR_COLUMN}{END_LSR_ROW}']):
        for j, cell in enumerate(row):
            lsr_sheet.cell(row=i+1+LSR_SHIFT, column=j+1, value=cell.value)
            lsr_sheet.cell(row=i+1+LSR_SHIFT, column=j+1).font = copy(cell.font)
            lsr_sheet.cell(row=i+1+LSR_SHIFT, column=j+1).border = copy(cell.border)
            lsr_sheet.cell(row=i+1+LSR_SHIFT, column=j+1).fill = copy(cell.fill)
            lsr_sheet.cell(row=i+1+LSR_SHIFT, column=j+1).number_format = copy(cell.number_format)
            lsr_sheet.cell(row=i+1+LSR_SHIFT, column=j+1).alignment = copy(cell.alignment)
    
    for i, row in enumerate(Main_Sheet[f"{START_SOURCE_COLUMN}{START_SOURCE_ROW}:{END_SOURCE_COLUMN}{SOURCE_NUMS_ROW}"]):
        for j, cell in enumerate(row):
            if i+1==int(SOURCE_NUMS_ROW): lsr_sheet.cell(row=i+1+SOURCE_SHIFT, 
                                    column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN), 
                                    value=int(cell.value)+xl.utils.cell.column_index_from_string(END_LSR_COLUMN))
            else: lsr_sheet.cell(row=i+1+SOURCE_SHIFT, 
                                 column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN), 
                                 value=cell.value)
            lsr_sheet.cell(row=i+1+SOURCE_SHIFT, column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).font = copy(cell.font)
            lsr_sheet.cell(row=i+1+SOURCE_SHIFT, column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).border = copy(cell.border)
            lsr_sheet.cell(row=i+1+SOURCE_SHIFT, column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).fill = copy(cell.fill)
            lsr_sheet.cell(row=i+1+SOURCE_SHIFT, column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).number_format = copy(cell.number_format)
            lsr_sheet.cell(row=i+1+SOURCE_SHIFT, column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).protection = copy(cell.protection)
            lsr_sheet.cell(row=i+1+SOURCE_SHIFT, column=j+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).alignment = copy(cell.alignment)
    
    # Копирование объединений ячеек
    for elem in lsr_for_copy.merged_cells:
        start_coord, end_coord = str(elem).split(':')
        # Проверка, чтобы начало объединения было выше и левее конца считываемой области, иначе нет смысла его переносить
        if xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string(start_coord)[0])<xl.utils.cell.column_index_from_string(END_LSR_COLUMN) and xl.utils.cell.coordinate_from_string(start_coord)[1]<int(END_LSR_ROW):
            # Проверка, чтобы объединение не выходило за считываемую область, иначе обрезается
            elem = [xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string(start_coord)[0]),
                    xl.utils.cell.coordinate_from_string(start_coord)[1],
                    xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string(end_coord)[0]),
                    xl.utils.cell.coordinate_from_string(end_coord)[1]]
            
            if elem[2]>xl.utils.cell.column_index_from_string(END_LSR_COLUMN):
                elem[2]=xl.utils.cell.column_index_from_string(END_LSR_COLUMN)
            if elem[3]>int(END_LSR_ROW):
                elem[3]=int(END_LSR_ROW)
            elem[0]=xl.utils.cell.get_column_letter(elem[0])
            elem[1]+=LSR_SHIFT
            elem[2]=xl.utils.cell.get_column_letter(elem[2])
            elem[3]+=LSR_SHIFT
            lsr_sheet.merge_cells(elem[0]+str(elem[1])+':'+elem[2]+str(elem[3]))
    
    for elem in Main_Sheet.merged_cells:
        start_coord, end_coord = str(elem).split(':')
        # Проверка, чтобы начало объединения было выше и левее конца считываемой области, иначе нет смысла его переносить
        if xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string(start_coord)[0])<xl.utils.cell.column_index_from_string(END_SOURCE_COLUMN) and xl.utils.cell.coordinate_from_string(start_coord)[1]<int(SOURCE_NUMS_ROW):
            # Проверка, чтобы объединение не выходило за считываемую область, иначе обрезается
            elem = [xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string(start_coord)[0]),
                    xl.utils.cell.coordinate_from_string(start_coord)[1],
                    xl.utils.cell.column_index_from_string(xl.utils.cell.coordinate_from_string(end_coord)[0]),
                    xl.utils.cell.coordinate_from_string(end_coord)[1]]
            
            if elem[2]>xl.utils.cell.column_index_from_string(END_SOURCE_COLUMN):
                elem[2]=xl.utils.cell.column_index_from_string(END_SOURCE_COLUMN)
            if elem[3]>int(SOURCE_NUMS_ROW):
                elem[3]=int(SOURCE_NUMS_ROW)
            elem[0]=xl.utils.cell.get_column_letter(elem[0]+xl.utils.cell.column_index_from_string(END_LSR_COLUMN))
            elem[1]+=SOURCE_SHIFT
            elem[2]=xl.utils.cell.get_column_letter(elem[2]+xl.utils.cell.column_index_from_string(END_LSR_COLUMN))
            elem[3]+=SOURCE_SHIFT
            lsr_sheet.merge_cells(elem[0]+str(elem[1])+':'+elem[2]+str(elem[3]))

    # Копирование ширины столбцов и высоты строк и скрытие лишних столбцов
    for column in range(xl.utils.cell.column_index_from_string(START_LSR_COLUMN), xl.utils.cell.column_index_from_string(END_LSR_COLUMN)+1):
            lsr_sheet.column_dimensions[get_column_letter(column)].width = int(lsr_for_copy.column_dimensions[get_column_letter(column)].width)
            lsr_sheet.column_dimensions[get_column_letter(column)].bestFit = True
    
    for column in range(xl.utils.cell.column_index_from_string(START_SOURCE_COLUMN), xl.utils.cell.column_index_from_string(END_SOURCE_COLUMN)+1):
            lsr_sheet.column_dimensions[get_column_letter(column+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN))].width = int(Main_Sheet.column_dimensions[get_column_letter(column+1)].width)

    for row in range(int(START_LSR_ROW), int(END_LSR_ROW)):
            lsr_sheet.row_dimensions[row+LSR_SHIFT].height = lsr_for_copy.row_dimensions[row].height
    
    lsr_sheet.column_dimensions.group(get_column_letter(xl.utils.cell.column_index_from_string(END_LSR_INFO_COLUMN)+1), END_LSR_COLUMN, hidden=True)
    lsr_sheet.row_dimensions.group(int(END_LSR_INFO_ROW), int(END_LSR_ROW), hidden=True)

    # Добавление гиперссылки на скопированный лист
    add_hyperlink_row(Hyperlinks_Worksheet, Main_Sheet, lsr_sheet.title, lsr_document[1], pasted_links)


    # Добавление заголовка с названием листа в список ошибок
    pasted_errors = add_error_name_header(Errors_Worksheet, Main_Sheet, lsr_sheet.title, pasted_errors)

    # Поиск и сопоставление позиций в ЛСР
    # Перебор всех искомых строк в ЛСР
    for a, lsr_row in enumerate(lsr_sheet[f"{START_LSR_COLUMN}{int(LSR_NUMS_ROW)+int(LSR_SHIFT)+1}:{END_LSR_INFO_COLUMN}{END_LSR_INFO_ROW}"]):
        # Проверка на пустую строку
        if lsr_row[xl.utils.cell.column_index_from_string(LSR_NAME_COLUMN)-1].value != None and lsr_row[xl.utils.cell.column_index_from_string(LSR_UNIT_COLUMN)-1].value != None:
            # Перебор всех строк в файлах - источниках
            parse_state=False
            for source_sheet in source_sheets:
                source_sheet = RES_FILE[source_sheet]
                for b, source_row in enumerate(source_sheet[f"{START_SOURCE_COLUMN}{int(SOURCE_NUMS_ROW)+int(SOURCE_SHIFT)+1}:{END_SOURCE_COLUMN}{END_SOURCE_ROW}"]):
                    EMPTY_SOURCE_MARKER = False
                    for cell in source_row:
                        if cell.value in [None, 'None', '']:
                            EMPTY_SOURCE_MARKER = True
                    
                    # Если есть полное совпадение без пустых ячеек
                    if (lsr_row[xl.utils.cell.column_index_from_string(LSR_NAME_COLUMN)-1].value == source_row[xl.utils.cell.column_index_from_string(SOURCE_NAME_COLUMN)-1].value and 
                        lsr_row[xl.utils.cell.column_index_from_string(LSR_UNIT_COLUMN)-1].value == source_row[xl.utils.cell.column_index_from_string(SOURCE_UNIT_COLUMN)-1].value and
                        EMPTY_SOURCE_MARKER==False):
                        for n, source_cell in enumerate(source_row):
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).value = copy(source_cell.value)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).font = copy(source_cell.font)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).border = copy(source_cell.border)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).fill = copy(source_cell.fill)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).number_format = copy(source_cell.number_format)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).alignment = copy(source_cell.alignment)
                        parse_state = True
                        break

                    # Если есть полное совпадение с пустыми ячейками
                    if (lsr_row[xl.utils.cell.column_index_from_string(LSR_NAME_COLUMN)-1].value == source_row[xl.utils.cell.column_index_from_string(SOURCE_NAME_COLUMN)-1].value and 
                        lsr_row[xl.utils.cell.column_index_from_string(LSR_UNIT_COLUMN)-1].value == source_row[xl.utils.cell.column_index_from_string(SOURCE_UNIT_COLUMN)-1].value and
                        EMPTY_SOURCE_MARKER==True):
                        for n, source_cell in enumerate(source_row):
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).value = copy(source_cell.value)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).font = copy(source_cell.font)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).border = copy(source_cell.border)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).fill = copy(source_cell.fill)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).number_format = copy(source_cell.number_format)
                            lsr_sheet.cell(row=a+1+int(LSR_NUMS_ROW)+int(LSR_SHIFT), column=n+1+xl.utils.cell.column_index_from_string(END_LSR_COLUMN)).alignment = copy(source_cell.alignment)
                        pasted_errors = add_error_row(Errors_Worksheet, Main_Sheet, pasted_errors, lsr_row, 103, source_row, source_sheet.title, b+int(SOURCE_NUMS_ROW)+int(SOURCE_SHIFT)+1)
                        parse_state = True
                        break
                    
                    # Если есть несовпадение по единице измерения
                    if (lsr_row[xl.utils.cell.column_index_from_string(LSR_NAME_COLUMN)-1].value == source_row[xl.utils.cell.column_index_from_string(SOURCE_NAME_COLUMN)-1].value and 
                        lsr_row[xl.utils.cell.column_index_from_string(LSR_UNIT_COLUMN)-1].value != source_row[xl.utils.cell.column_index_from_string(SOURCE_UNIT_COLUMN)-1].value):
                        pasted_errors = add_error_row(Errors_Worksheet, Main_Sheet, pasted_errors, lsr_row, 104, source_row, source_sheet.title, b+int(SOURCE_NUMS_ROW)+int(SOURCE_SHIFT)+1)
                        parse_state = True
                        break
                
                if parse_state: break
            if not parse_state:
                pasted_errors = add_error_row(Errors_Worksheet, Main_Sheet, pasted_errors, lsr_row, 101)


    return RES_FILE, pasted_errors


def insert_info(source_path, source_sheets, lsr_documents_states) -> None:
    """
    Заолнение выбранных файлов найденными материалами

    :return: None
    """
    # Книга, в которую будет записана вся итоговая информация
    RES_FILE = xl.load_workbook(source_path)

    # Первый лист с источником информации для копирования шапки таблицы и т.д.
    Main_Sheet = RES_FILE[source_sheets[0]]

    # Лист с гиперссылками на обработанные ЛСР в итоговой книге
    Hyperlinks_Worksheet = RES_FILE.create_sheet("Гипрессылки", 0)
    pasted_links = create_hyperlinks_table_header(Hyperlinks_Worksheet, Main_Sheet, source_sheets)

    # Лист с ошибками из ЛСР в итоговой книге
    Errors_Worksheet = RES_FILE.create_sheet("Ошибки", 0)
    pasted_errors = create_error_table_header(Errors_Worksheet, Main_Sheet)
    
    # Обработка каждого файла ЛСР в отдельности
    for num, lsr_document in enumerate(lsr_documents_states):
        # Полное копирование ЛСР в итоговую книгу
        RES_FILE, pasted_errors = add_lsr_sheet(RES_FILE, lsr_document, Hyperlinks_Worksheet, Errors_Worksheet, source_sheets, num+pasted_links+1, pasted_errors)
    
    '''
        add_error_name_header(error_worksheet, main_sheet, sheet, pasted_error_rows)
        pasted_error_rows+=1

        for row in worksheet[f"C5:C{worksheet.max_row}"]:
            parse_coordinate=0
            for cell in row:
                try:
                    if len(str(cell.value))>=3:
                        parse_coordinate=cell.row
                except Exception as e: pass
            if parse_coordinate: break


        for a, row in enumerate(worksheet[f"A{parse_coordinate}:E{worksheet.max_row}"]):
            bufer_data = None
            error_code = 101
            if row[2].value!=None and row[3].value!=None:
                for parse_sheet in sheets[0]:
                    parse_sheet = file[parse_sheet]
                    
                    for b, parse_row in enumerate(parse_sheet[f"D5:AV{parse_sheet.max_row}"]):
                        # если информация полностью совпадает и все цены есть
                        if str(row[2].value)==str(parse_row[0].value) and str(row[3].value)==str(parse_row[1].value) and str(parse_row[11].value)!=None and str(parse_row[25].value)!=None and str(parse_row[39].value)!=None:
                            for copy_row in parse_sheet[f"B{b+5}:AV{b+5}"]:
                                for i, copy_cell in enumerate(copy_row):
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).value = copy(copy_cell.value)
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).font = copy(copy_cell.font)
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).border = copy(copy_cell.border)
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).fill = copy(copy_cell.fill)
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).number_format = copy(copy_cell.number_format)
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).protection = copy(copy_cell.protection)
                                    worksheet.cell(row=a+parse_coordinate, column=i+start_coordinate[1]).alignment = copy(copy_cell.alignment)
                            bufer_data = None
                            parse_state = True
                            break
                        # если нет цены 
                        if str(row[2].value)==str(parse_row[0].value) and str(row[3].value)==str(parse_row[1].value) and (str(parse_row[11].value)==None or str(parse_row[25].value)==None or str(parse_row[39].value)==None):
                            bufer_data = parse_sheet[f"A{b+5}:AV{b+5}"][0]
                            error_code = 102
                            parse_state=False
                        #если не совпадает ед. изм.
                        if str(row[2].value)==str(parse_row[0].value) and str(row[3].value)!=str(parse_row[1].value):
                            bufer_data = parse_sheet[f"A{b+5}:AV{b+5}"][0]
                            error_code = 103
                            parse_state=False
                        else:
                            parse_state=False
                    if parse_state: break
                
                if not parse_state:
                    add_error_row(error_worksheet, main_sheet, pasted_error_rows, row, error_code, bufer_data)
                    pasted_error_rows+=1
    '''
    RES_FILE.save(source_path[:-5]+" Обработанный "+source_path[-5:])
    return None

if __name__ == "__main__":
    insert_info('/Users/Dima/Desktop/Общий (рабочий) 21.8.xlsx', 
                ['КА Материалы 23.6.23', 'КА Оборудование 27.6.23'],
                [['ЛСР-1-02-02-02', 'изм.1 Силовые кабели 0,4 кВ. КЖ (027-ЭП2.2) - Ведомость объемов работ по смете',
                '/Users/Dima/Desktop/ЛСР/ЛСР-1-02-02-02 изм.1 Силовые кабели 0,4 кВ. КЖ (027-ЭП2.2) - Ведомость объемов работ по смете.xlsx'],
                ['ЛСР-1-02-02-01', 'изм.1 Высоковольтные кабели. КЖ (027-ЭП1.2) - Ведомость объемов работ по смете',
                '/Users/Dima/Desktop/ЛСР/ЛСР-1-02-02-01 изм.1 Высоковольтные кабели. КЖ (027-ЭП1.2) - Ведомость объемов работ по смете.xlsx'],
                ['ЛСР-1-02-01-01', 'изм.1 ОРУ 220кВ. ЭП (113-ЭП)_ ДО - Ведомость объемов работ по смете',
                '/Users/Dima/Desktop/ЛСР/ЛСР-1-02-01-01 изм.1 ОРУ 220кВ. ЭП (113-ЭП)_ ДО - Ведомость объемов работ по смете.xlsx'],
                ['ЛСР-1-02-02-03', 'изм.1 Кабельное хозяйство. ЭП (027-ЭП5.3) - Ведомость объемов работ по смете',
                '/Users/Dima/Desktop/ЛСР/ЛСР-1-02-02-03 изм.1 Кабельное хозяйство. ЭП (027-ЭП5.3) - Ведомость объемов работ по смете.xlsx'],
                ['ЛСР-1-02-03-01', 'изм.1 ЗРУ 110 кВ с ЗРУ 35 кВ. ЭП (127-ЭП) - Ведомость объемов работ по смете',
                '/Users/Dima/Desktop/ЛСР/ЛСР-1-02-03-01 изм.1 ЗРУ 110 кВ с ЗРУ 35 кВ. ЭП (127-ЭП) - Ведомость объемов работ по смете.xlsx'],
                ['ЛСР-1-02-01-02', 'изм.1 ОРУ 110кВ. ЭП (113-ЭП1)_ ДО - Ведомость объемов работ по смете',
                '/Users/Dima/Desktop/ЛСР/ЛСР-1-02-01-02 изм.1 ОРУ 110кВ. ЭП (113-ЭП1)_ ДО - Ведомость объемов работ по смете.xlsx']])